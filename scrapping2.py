import requests
from bs4 import BeautifulSoup
import json
import pandas as pd
from openpyxl import load_workbook

# URL do site de onde queremos extrair os dados
url = "https://www.bagaggio.com.br/malas?order=OrderByReleaseDateDESC"

# Fazer a solicitação HTTP para obter o conteúdo da página
response = requests.get(url)
html_content = response.content

# Parsing do conteúdo HTML usando BeautifulSoup
soup = BeautifulSoup(html_content, 'html.parser')

# Localizar o script JSON que contém as informações do produto
script_tag = soup.find('script', string=lambda t: t and 'Product:sp-' in t)
if not script_tag:
    print("Tag de script com informações do produto não encontrada")
    exit()

# Extrair o texto JSON do script
json_text = script_tag.string.strip()
json_text = json_text[json_text.find("{"):json_text.rfind("}")+1]

# Ajustar o texto JSON para garantir que esteja correto
json_text = json_text.replace('undefined', 'null')

# Carregar o JSON
try:
    json_data = json.loads(json_text)
except json.JSONDecodeError as e:
    print(f"Erro ao carregar JSON: {e}")
    exit()

# Lista para armazenar informações dos produtos
product_data = []

# Função para extrair valores de propriedades
def get_property_value(properties, prop_name):
    for prop in properties:
        if prop.get('name') == prop_name:
            return prop.get('values', {}).get('json', [None])[0]
    return None

# Iterar sobre os produtos no JSON
for key, value in json_data.items():
    if key.startswith("Product:sp-"):
        properties = value.get("properties", [])
        items = value.get('items({"filter":"ALL_AVAILABLE"})', [])
        for item in items:
            product_info = {
                "Nome": item.get("nameComplete"),
                "Item ID": item.get("itemId"),
                "EAN": item.get("ean"),
                "Marca": get_property_value(properties, "Marca"),
                "Suporta Até": get_property_value(properties, "Suporta Até"),
                "Peso": get_property_value(properties, "Peso (kg)"),
                "Garantia": get_property_value(properties, "Garantia"),
                "Idade": get_property_value(properties, "Idade"),
                "Capacidade": get_property_value(properties, "Capacidade (L)"),
                "Dimensões": get_property_value(properties, "Dimensões"),
                "Cor": get_property_value(properties, "Cor"),
                "Vai à Bordo": get_property_value(properties, "Vai à bordo (Comporta até 8kg)"),
                "Cadeado com Senha": get_property_value(properties, "Cadeado com Senha"),
                "Rodas 360º": get_property_value(properties, "Rodas 360º"),
                "Alça de Mão Superior": get_property_value(properties, "Alça de mão superior"),
                "Material": get_property_value(properties, "Mala de ABS"),
                "Tamanho": get_property_value(properties, "Tamanho"),
                "Seller ID": get_property_value(properties, "sellerId"),
            }
            product_data.append(product_info)

# Converter os dados para um DataFrame do Pandas
df = pd.DataFrame(product_data)

# Escrever os dados em um arquivo Excel temporário usando pandas
temp_excel_path = r"C:\Users\BG-PROVISORIO\Desktop\temp_teste.xlsx"
df.to_excel(temp_excel_path, index=False)

# Carregar o arquivo Excel com openpyxl para manipular a posição dos dados
wb = load_workbook(temp_excel_path)
ws = wb.active

# Defina a célula inicial onde deseja começar a escrever os dados
start_row = 10
start_col = 5

# Escrever os cabeçalhos
for col_num, column_title in enumerate(df.columns, start=start_col):
    ws.cell(row=start_row, column=col_num, value=column_title)

# Escrever os dados
for row_num, row_data in enumerate(df.values, start=start_row + 1):
    for col_num, cell_value in enumerate(row_data, start=start_col):
        ws.cell(row=row_num, column=col_num, value=cell_value)

# Salvar o arquivo Excel final
final_excel_path = r"C:\Users\BG-PROVISORIO\Desktop\teste.xlsx"
wb.save(final_excel_path)

print("Dados extraídos e salvos com sucesso!")
